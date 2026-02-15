import json
from itertools import combinations
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule

KANJI_PATH = "data/kanji_data_1.json"
RADICAL_PATH = "data/radical_data.json"
OUTPUT_XLSX = "radical_triplets.xlsx"

# ------------------------------------------------------------
# Load data
# ------------------------------------------------------------
with open(KANJI_PATH, "r", encoding="utf-8") as f:
    kanji_data = json.load(f)

with open(RADICAL_PATH, "r", encoding="utf-8") as f:
    radical_info = json.load(f)

# ------------------------------------------------------------
# Radical metadata
# ------------------------------------------------------------
radical_numbers = sorted(int(k) for k in radical_info.keys())
radical_chars = {int(k): radical_info[k]["radical_char"] for k in radical_info}
radical_labels = {r: f"{r}. {radical_chars[r]}" for r in radical_numbers}

# ------------------------------------------------------------
# Count triplets with full permutation symmetry
# ------------------------------------------------------------
triplet_counts = {}

for entry in kanji_data.values():
    rads = sorted(set(entry.get("radicals", [])))

    # True triplets (three distinct radicals)
    for a, b, c in combinations(rads, 3):
        perms = [
            (a, b, c), (a, c, b),
            (b, a, c), (b, c, a),
            (c, a, b), (c, b, a)
        ]
        for key in perms:
            triplet_counts[key] = triplet_counts.get(key, 0) + 1

    # Pair case (two distinct radicals, third repeated)
    for a, b in combinations(rads, 2):
        perms = [
            (a, b, b),
            (b, a, b),
            (b, b, a)
        ]
        for key in perms:
            triplet_counts[key] = triplet_counts.get(key, 0) + 1

# ------------------------------------------------------------
# Build workbook
# ------------------------------------------------------------
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

sheet_scores = []  # (score, radical_a, sheet)

for a in radical_numbers:

    # Collect all b,c that appear with a
    related = set()
    for (x, b, c), v in triplet_counts.items():
        if x == a and v > 0:
            related.add(b)
            related.add(c)

    if not related:
        continue

    related = sorted(related)

    # Build matrix
    matrix = {b: {c: 0 for c in related} for b in related}

    for (x, b, c), v in triplet_counts.items():
        if x == a and b in related and c in related:
            matrix[b][c] = v

    # Remove rows/columns that are all zero
    nonzero = []
    for b in related:
        if any(matrix[b][c] != 0 for c in related):
            nonzero.append(b)

    related = nonzero
    if not related:
        continue

    # Sort by diagonal descending
    related.sort(key=lambda r: matrix[r][r], reverse=True)

    # Create sheet
    ws = wb.create_sheet(title=radical_labels[a])

    # Header row
    ws.append([""] + [radical_labels[b] for b in related])

    # Rows
    for b in related:
        ws.append([radical_labels[b]] + [matrix[b][c] for c in related])

    # Only apply formatting if we have at least a 1×1 data region
    if len(related) >= 1:
        data_start_row = 2
        data_start_col = 2
        data_end_row = 1 + len(related)
        data_end_col = 1 + len(related)

        # Ensure the range is valid (end must be >= start)
        if data_end_row >= data_start_row and data_end_col >= data_start_col:
            cell_range = (
                f"{ws.cell(row=data_start_row, column=data_start_col).coordinate}:"
                f"{ws.cell(row=data_end_row, column=data_end_col).coordinate}"
            )

            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="FFFFFF",
                end_type="num", end_value=100, end_color="00B050"
            )

            ws.conditional_formatting.add(cell_range, rule)

    # Score for sheet ordering
    top_left = matrix[related[0]][related[0]]
    sheet_scores.append((top_left, a, ws))

# ------------------------------------------------------------
# Sort sheets by top-left diagonal value
# ------------------------------------------------------------
sheet_scores.sort(key=lambda x: x[0], reverse=True)

# Reorder sheets
for i, (_, _, ws) in enumerate(sheet_scores):
    wb._sheets.remove(ws)
    wb._sheets.insert(i, ws)

# ------------------------------------------------------------
# Save workbook
# ------------------------------------------------------------
wb.save(OUTPUT_XLSX)
